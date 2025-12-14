from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook


@dataclass
class SheetContent:
    name: str
    cells: List[List[str]]  # 2D grid (rows x cols)


@dataclass
class XlsxContent:
    sheets: List[SheetContent]


def read_xlsx_content(path: Path, max_rows: int = 200, max_cols: int = 50) -> XlsxContent:
    """Read a bounded grid of values from each sheet."""
    wb = load_workbook(str(path), data_only=True, read_only=True)
    sheets: List[SheetContent] = []

    for ws in wb.worksheets:
        grid: List[List[str]] = []
        for r_i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True), start=1):
            grid.append(["" if v is None else str(v) for v in row])
        sheets.append(SheetContent(name=ws.title, cells=grid))

    return XlsxContent(sheets=sheets)
